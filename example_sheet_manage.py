from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() #새로운 시트 생성

ws.title = 'Mysheet'
ws.sheet_properties.tabColor = 'ff66ff'

ws1 = wb.create_sheet('YourSheet') # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet('NewSheet',2) # 시트 index 2 번째에 Sheet 생성

# 시트 접근은 ws1, ws2, ..처럼 하는 방법도 있고, wb['시트명'] 처럼 dict 형태로도 접근 가능
print(wb['NewSheet'].title)
new_ws = wb['NewSheet']
print(new_ws.title)

#모든 시트 확인, 리스트로 변환

print(wb.sheetnames)

# Sheet 복사
new_ws['A1'] = 'TEST' # A1셀에 데이터 넣음

target = wb.copy_worksheet(new_ws) # 복사된 Sheet가 우측 마지막에 생성됨(데이터포함)
target.title = 'Copied_Sheet'

wb.save('nadoCoding_sample2.xlsx')