import openpyxl
from openpyxl import workbook
from random import *

wb = openpyxl.load_workbook(filename= 'nadoCoding_sample.xlsx')
ws = wb.active

# 셀에 데이터(값) 입력

ws['A1'] = 1
ws['A2'] = 2
ws['B1'] = 3
ws['B2'] = 4


print(ws['A1']) # <Cell 'Sheet'.A1> - 셀 객체 정보만 출력
print(ws['A1'].value) # A1- 입력된 값만 출력
print(ws['A1'].value) # 값이 없을때는 'None' 이 출력


#엑셀에서 행(row) = 1, 2, 3, .... /열(colum()은 A, B, C, ... 열에 대해 1, 2, 3 지칭가능


print(ws.cell(row=1, column=1).value)
print(ws.cell(1, 1).value)

ws.cell(1, 3).value = 10
ws.cell(2, 3, value=20)
c = ws.cell(3, 3, value=30)


print(c.value)

#반복문으로 랜덤 숫자 채워보기

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(x, y).value = randint(0, 100)


wb.save('cell_manager_example.xlsx')